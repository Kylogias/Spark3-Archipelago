import openpyxl as xl
from openpyxl.styles.colors import *
from openpyxl.styles.fills import *
from openpyxl.styles.alignment import *
from openpyxl.styles.borders import *
from enum import Enum
from logicshared import shared
import time

class RuleToken(Enum):
	UNK = 0
	LPAREN  = 1
	RPAREN  = 2
	AND     = 3
	OR      = 4
	FALSE   = 5

STRING_TO_TOKEN = [
	" ", "(", ")", "+", "|"
]

class And:
	def __init__(self, a, b):
		self.options = []
		self.add_to_opt(a)
		self.add_to_opt(b)
	def add_to_opt(self, opt):
		if isinstance(opt, And):
			for option in opt.options:
				self.options.append(option)
		elif opt == True or opt == None:
			pass
		else:
			self.options.append(opt)
	def __str__(self):
		return self.__repl__()
	def __repl__(self):
		string = "("
		for option in self.options:
			string += str(option)
			if option != self.options[-1]:
				string += " + "
		string += ")"
		return string
	def resolve(self, collection):
		ors = []
		for option in self.options:
			if isinstance(option, Or):
				ors.append(option)
			else:
				for collect in collection:
					if not option in collect:
						collect.append(option)
		for option in ors:
			option.resolve(collection)
	def can_complete(self, collect):
		for option in self.options:
			if isinstance(option, str):
				if not option in collect: return False
			if isinstance(option, Or):
				if not option.can_complete(collect): return False
		return True

class Or:
	def __init__(self, a, b):
		self.options = []
		self.always_true = False
		self.add_to_opt(a)
		self.add_to_opt(b)
		if self.always_true:
			self.options = []
	def add_to_opt(self, opt):
		if isinstance(opt, Or):
			for option in opt.options:
				self.options.append(option)
		elif opt == True or opt == None:
			self.always_true = True
		else:
			self.options.append(opt)
	def __str__(self):
		return self.__repl__()
	def __repl__(self):
		string = "("
		for option in self.options:
			string += str(option)
			if option != self.options[-1]:
				string += " | "
		string += ")"
		return string
	def resolve(self, collection):
		for collect in collection.copy():
			if not self.can_complete(collect):
				collection.remove(collect)
				for option in self.options:
					if isinstance(option, str):
						collection.append(collect.copy() + [option])
					elif isinstance(option, And):
						new_collect = [collect.copy()]
						option.resolve(new_collect)
						for new in new_collect:
							collection.append(new)
	def can_complete(self, collect):
		for option in self.options:
			if isinstance(option, str):
				if option in collect: return True
			if isinstance(option, And):
				if option.can_complete(collect): return True
		return False

def add_to_rule(op, rule, add):
	if op == RuleToken.UNK: return add
	if op == RuleToken.AND: return And(rule, add)
	if op == RuleToken.OR: return Or(rule, add)

def recurse_tokens(tokens):
	op = RuleToken.UNK
	rule = None
	while len(tokens):
		token = tokens.pop(0)
		if isinstance(token, RuleToken):
			match token:
				case RuleToken.UNK: pass
				case RuleToken.LPAREN: rule = add_to_rule(op, rule, recurse_tokens(tokens))
				case RuleToken.RPAREN: return rule
				case RuleToken.AND: op = RuleToken.AND
				case RuleToken.OR: op = RuleToken.OR
		else:
			rule = add_to_rule(op, rule, token)
			
	if rule == None:
		rule = True
	return rule

def parse_rule_string(rule):
	tokens = []
	i = 0
	while i < len(rule):
		tok_added = False
		for tok in range(len(STRING_TO_TOKEN)):
			if rule[i:].startswith(STRING_TO_TOKEN[tok]):
				tokens.append(RuleToken(tok))
				i += len(STRING_TO_TOKEN[tok])
				tok_added = True
				break
		if not tok_added:
			for item in shared["items"]:
				if "rule" in item and rule[i:].startswith(item["rule"]):
					tok_added = True
					i += len(item["rule"])
					tokens.append(item["name"])
					break
		if not tok_added:
			for macro in shared["rule_macros"]:
				if rule[i:].startswith(macro["rule"]):
					tokens += parse_rule_string(macro["expansion"])
					i += len(macro["rule"])
					tok_added = True
					break
	return tokens

def recurse_region_rule(stage, region, seen):
	rule = {}
	for difficulty in shared["difficulties"]:
		rule[difficulty] = ""
	entrances = []
	for ent in region["entrances"]:
		if ent in seen: continue
		entrance_rule = region["entrances"][ent]
		for entrance in stage["regions"]:
			if entrance["name"] == ent:
				recursed_seen = seen.copy()
				recursed_seen.append(ent)
				recursive_rule = recurse_region_rule(stage, entrance, recursed_seen)
				for difficulty in recursive_rule:
					entrances.append(f"(({recursive_rule[difficulty]})+({entrance_rule[difficulty if difficulty in entrance_rule else 'base']}))")
	for difficulty in rule:
		all_entrances = ""
		for entrance in entrances:
			if entrance == entrances[0]:
				all_entrances = f"{entrance}"
			else:
				all_entrances = f"{all_entrances}|{entrance}"
				
		rule[difficulty] = f"({all_entrances})"
	return rule

colors = {"region_header": ["d090d0"], "check_header": ["7070d0", "6767c7"], "base": ["a0d0a0", "97c797"], "hard": ["d07070", "c76767"]}
class SheetState:
	def __init__(self, ws, style, color_index):
		self.column_size = 6
		self.start_row = 1
		self.current_row = 1
		self.current_column = 1
		self.max_row = 1
		self.ws = ws
		self.color = PatternFill("solid", fgColor=Color(colors[style][color_index % len(colors[style])]))
	def next_row(self, new_style=None, new_color=0):
		if self.current_row > self.max_row: self.max_row = self.current_row
		self.current_row += 1
		if new_style != None:
			self.color = PatternFill("solid", fgColor=Color(colors[new_style][color_index % len(colors[new_style])]))
		return self
	def next_column(self, new_style=None, new_color=0):
		self.current_column += self.column_size
		if new_style != None:
			self.color = PatternFill("solid", fgColor=Color(colors[new_style][color_index % len(colors[new_style])]))
		return self
	def reset_to_next_row(self, new_style=None, new_color=0):
		self.current_column = 1
		self.current_row = self.max_row+1
		self.start_row = self.current_row
		self.max_row = self.current_row
		if new_style != None:
			self.color = PatternFill("solid", fgColor=Color(colors[new_style][color_index % len(colors[new_style])]))
		return self
	def reset_to_next_column(self, new_style=None, new_color=0):
		self.current_column += self.column_size
		self.current_row = self.start_row
		if new_style != None:
			self.color = PatternFill("solid", fgColor=Color(colors[new_style][color_index % len(colors[new_style])]))
		return self
	def add_cell(self, text, merge_right=False, center_horizontal=False):
		c = self.ws.cell(self.current_row, self.current_column, text)
		c.fill = self.color
		if merge_right:
			self.ws.merge_cells(start_row=self.current_row, start_column=self.current_column, end_row=self.current_row, end_column=self.current_column+self.column_size-1)
		if center_horizontal:
			c.alignment = Alignment(horizontal="center")
			side = Side("thin", Color("000000"))
			c.border = Border(left=side, right=side, top=side, bottom=side)
		return self
	def merge_column_down(self, column):
		merge_column = column*self.column_size+1
		self.ws.merge_cells(start_row=self.start_row, start_column=merge_column, end_row=self.max_row, end_column=merge_column+self.column_size-1)
		c = self.ws.cell(self.start_row, merge_column)
		c.alignment = Alignment(horizontal="center", vertical="center")
		side = Side("thin", Color("000000"))
		c.border = Border(left=side, right=side)
		return self

def make_path_strings(rule):
	paths = []
	resolved = recurse_tokens(parse_rule_string(rule))
	if isinstance(resolved, And) or isinstance(resolved, Or):
		temp = [[]]
		resolved.resolve(temp)
		resolved = temp
	if resolved == True:
		return [""]
	if isinstance(resolved, str):
		return [resolved]
	for collect in resolved:
		string = ""
		for item in collect:
			string += item
			if item != collect[-1]:
				string += ", "
		paths.append(string)
	return paths

wb = xl.Workbook()
ws = wb.active
rules = []
for item in shared["items"]:
	if "rule" in item:
		rules.append(item["name"])
for stage in shared["stages"]:
	ws = wb.create_sheet(title=stage["name"])
	sheet = SheetState(ws, "region_header", 0)
	sheet.add_cell("Region", True, True).next_column().add_cell("Entrance", True, True).next_column()
	for difficulty in shared["difficulties"]:
		sheet.add_cell(f"{difficulty} difficulty", True, True).next_column()
	color_index = 0
	sheet.reset_to_next_row("check_header", color_index)
	stage_checks = []
	for region in stage["regions"]:
		for entrance in region["entrances"]:
			sheet.add_cell(region["name"]).next_column().add_cell(entrance)
			for difficulty in shared["difficulties"]:
				sheet.reset_to_next_column(difficulty, color_index)
				require = region["entrances"][entrance]
				paths = make_path_strings(require[difficulty if difficulty in require else "base"])
				for path in paths:
					sheet.add_cell(path, True).next_row()
			sheet.merge_column_down(0)
			sheet.merge_column_down(1)
			color_index += 1
			sheet.reset_to_next_row("check_header", color_index)
		for check in region["checks"]:
			stage_checks.append([check["name"], region["name"], check["requires"]])
	sheet.reset_to_next_row("region_header", 0)
	sheet.add_cell("Check", True, True).next_column().add_cell("Region", True, True).next_column()
	for difficulty in shared["difficulties"]:
		sheet.add_cell(f"{difficulty} difficulty", True, True).next_column()
	color_index = 0
	sheet.reset_to_next_row("check_header", color_index)
	for check in stage_checks:
		sheet.add_cell(check[0]).next_column().add_cell(check[1])
		for difficulty in shared["difficulties"]:
			sheet.reset_to_next_column(difficulty, color_index)
			require = check[2]
			paths = make_path_strings(require[difficulty if difficulty in require else "base"])
			for path in paths:
				sheet.add_cell(path, True).next_row()
		sheet.merge_column_down(0)
		sheet.merge_column_down(1)
		color_index += 1
		sheet.reset_to_next_row("check_header", color_index)

		'''
		region_rule = recurse_region_rule(stage, region, [])
		for check in region["checks"]:
			stage_checks.append([check["name"], region["name"], check["requires"]])
			for difficulty in shared["difficulties"]:
				check_rule = f"{region_rule[difficulty]}+({check['requires'][difficulty if difficulty in check['requires'] else 'base']})"
				rule = recurse_tokens(parse_rule_string(check_rule))
				if True and (isinstance(rule, And) or isinstance(rule, Or)):
					temp = [[]]
					rule.resolve(temp)
					rule = temp
				for collect in rule:
					string = ""
					for item in collect:
						string += item
						if item != collect[-1]:
							string += ", "
		'''
					

wb.save("logicdoc.xlsx")